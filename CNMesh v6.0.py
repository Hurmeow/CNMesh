import openpyxl
import os
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side

path = os.getcwd()  # создаем переменную с путем корневой папки

# если в каталоге имеется папка Excel, делаем ее новой корневой папкой
if os.path.isdir(path + '\Excel'):
    os.chdir(path + '\Excel')
    path = os.getcwd()
    print('Директория есть')
    print('Путь ' + os.getcwd())
# если в каталоге нет папки Excel, то создаем ее, и делаем ее корневой папкой
else:
    os.mkdir("Excel")
    os.chdir(path + '\Excel')
    path = os.getcwd()
    print('Папка Excel создана')
    print('Новый путь ' + os.getcwd())

###################################################################

############### БЛОК НАЗНАЧЕНИЯ ФАЙЛА С ДАННЫМИ ###################

###################################################################
name_data = str(input('Имя файла: '))
file_data = name_data + '.xlsx'  # вводим имя файла который нужно загрузить
print(file_data + ' загружен')  # печатаем в строке
wb_data_path = os.path.join(path, file_data)  # задаем путь к созданной книге wb_output_create
wb_data = openpyxl.load_workbook(wb_data_path)  # загружаем созданную ранее книгу
wsn = list(wb_data.sheetnames)  # получаем название листов
print(wsn)  # печатаем название листов
sheet_data = wb_data.worksheets[0]  # выбираем лист в позиции 0 (первый лист)

###################################################################

########### БЛОК НАЗНАЧЕНИЯ ФАЙЛА C ВЫХОДНЫМИ ДАННЫМИ #############

###################################################################

file_output = name_data + '_Calculate_N' + '.xlsx'  # имя файла для вывода
print(file_output + ' создан')  # печатаем в строке
wb_output_create = openpyxl.Workbook()  # создаем книгу в каталоге для последующего обращения к ней
wb_output_create.save(file_output)  # сохраняем книгу
wb_output_create.close()  # закрываем книгу
wb_output_output_path = os.path.join(path, file_output)  # задаем путь к созданной книге wb_output_create
wb_output = openpyxl.load_workbook(wb_output_output_path, read_only=False)  # загружаем созданную ранее книгу
sheet_output = wb_output.create_sheet('x', 0)  # создаем лист в позиции 0
sheet_output = wb_output.create_sheet('y', 1)  # создаем лист в позиции 1
sheet_output = wb_output.worksheets[0]  # выбираем лист в позиции 0 (первый лист)
wb_output.save(file_output)  # сохраняем книгу


# Переменные для оформления таблицы
myFill = PatternFill(start_color='e8fdfb', end_color='e8fdfb', fill_type='solid')
myFont = Font(b=True, size=12)
myAlignment = Alignment(horizontal='center')
myBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


g = 0
while g < 2:  # цикл для последовательной обработки листов книги
    sheet_data = wb_data.worksheets[g]  # выбираем лист в позиции g
    sheet_output = wb_output.worksheets[g]  # выбираем лист в позиции g

    max_row_data = sheet_data.max_row  # считываем количество строк с данными
    max_column_data = sheet_data.max_column  # считываем количество столбцов с данными

    print(name_data)  # название файла
    print('Лист ' + str(wb_data.sheetnames[g]))  # название листа

    length = int(input('Длина котлована: '))  # длина котлована
    start_x = -(length / 2)
    height = int(input('Глубина котлована: '))  # глубина котлована

    if length == 10 or length == 20:
        step_x = 1  # шаг по длине
        R_global = round(1.2, 2)  # радиус поиска ближайших точек

    elif length == 60 or length == 100 or length == 150:
        step_x = 5  # шаг по длине
        R_global = round(2, 2)  # радиус поиска ближайших точек

    if height == 5:
        height = height + 5
        step_y = 0.5  # шаг по глубине
        iter_height = int(height / step_y)  # число шагов итератора по длине for y in range(1, iter_height + 2)

    elif height == 10:
        height = height + 5
        step_y = 1  # шаг по глубине
        iter_height = int(height / step_y)  # число шагов итератора по длине for y in range(1, iter_height + 2)

    elif height == 15:
        height = height + 7
        step_y = 1.5  # шаг по глубине
        iter_height = int(height / step_y) + 1  # число шагов итератора по длине for y in range(1, iter_height + 2)

    elif height == 20:
        height = height + 10
        step_y = 2  # шаг по глубине
        iter_height = int(height / step_y)  # число шагов итератора по длине for y in range(1, iter_height + 2)

    iter_length = int(length / step_x)  # число шагов итератора по длине for x in range(1, iter_length + 2)
    N_z = 0
    d_point = []  # создаем временный массив для точек попавших в R < R_global

    if g == 0:
        column_X = 4  # принимаем строку где содержатся значения X
        column_Y = 6  # принимаем строку где содержатся значения Y
        column_Z = 8  # принимаем строку где содержатся значения Z
    elif g == 1:
        column_X = 5  # принимаем строку где содержатся значения X
        column_Y = 6  # принимаем строку где содержатся значения Y
        column_Z = 7  # принимаем строку где содержатся значения Z

    for x in range(1, iter_length + 2):  # итерация точек для N(N_x)
        N_x = round((start_x + step_x * x - step_x), 2)  # координата нормальной сетки по X

        for y in range(1, iter_height + 2):  # итерация точек для N(N_y)
            N_y = round(-(step_y * y - step_y), 2)  # координата нормальной сетки по Y
            list.clear(d_point)  # очищаем временный массив от предыдущего шага

            if N_y < -height:  # условие для последней координаты, если она превышает координату глубины котлована
                N_y = -height

            # Условие подбора точек с радиусом R < R_global с записью в массив d_point
            # первая строка в Excel ВСЕГДА ДОЛЖНА быть шапка со значениями!!!!!!!!!!!!

            for r in range(2, max_row_data + 1):
                m_x = sheet_data.cell(row=r, column=column_X).value  # считываем и записываем в переменную значение из Excel (x)
                m_y = sheet_data.cell(row=r, column=column_Y).value  # считываем и записываем в переменную значение из Excel (y)
                m_z = sheet_data.cell(row=r, column=column_Z).value  # считываем и записываем в переменную значение из Excel (z)
                # print(N_x, N_y)
                # print(m_x, m_y)
                R = ((N_x - m_x) ** 2 + (N_y - m_y) ** 2) ** 0.5  # определяем длину от точки N(N_x, N_y), до точки M(m_x, m_y)
                i_data = r  # номер строки с данными точки M(m_x, m_y, m_z). Для анализа, выявления багов

                if R < R_global:
                    '''
                    формирование временного массива d_point[] для точки N(N_x, N_y), 
                    для последующего определения точек ближайших к точке N(N_x, N_y)
                    '''
                    list.extend(d_point, [[m_x, m_y, m_z, R, i_data]])

            R1 = 10000  # радиус от точки N до M1 (первоначальное значение 10000 для входа в условия if)
            R2 = 10000  # радиус от точки N до M2 (первоначальное значение 10000 для входа в условия if)
            R3 = 10000  # радиус от точки N до M3 (первоначальное значение 10000 для входа в условия if)
            Per = 10000  # периметр треугольника образованного точками M1, M2, M3 (первоначальное значение 10000 для входа в условия if)

            # подбираем ближайшие точки M1, M2, M3
            for i in range(len(d_point)):  # точка M1
                for j in range(3, 4):  # для 3 значению списка в массиве d_point[]
                    r_1 = d_point[i][j]  # запись радиуса в переменную r_1 по индексу массива

                    for k in range(i + 1, len(d_point)):  # точка M2 (по строке)
                        for s in range(3, 4):  # для 3 значению списка в массиве d_point[]
                            r_2 = d_point[k][s]  # запись радиуса в переменную r_2 по индексу массива

                            for v in range(k + 1, len(d_point)):  # точка M3
                                for w in range(3, 4):  # для 3 значение списка в массиве d_point[]
                                    r_3 = d_point[v][w]  # запись радиуса в переменную r_3 по индексу массива
                                    u_z = None  # обнуление перемещений u_z
                                    a, b, c = None, None, None  # обнуление значений a, b, c

                                    # определяем периметр треугольника образованного точками M1, M2, M3 (для качественной оценки)
                                    perimeter = ((d_point[k][0] - d_point[i][0]) ** 2 + (d_point[k][1] - d_point[i][1]) ** 2) ** 0.5 + (
                                            (d_point[v][0] - d_point[k][0]) ** 2 + (d_point[v][1] - d_point[k][1]) ** 2) ** 0.5 + (
                                            (d_point[i][0] - d_point[v][0]) ** 2 + (d_point[i][1] - d_point[v][1]) ** 2) ** 0.5

                                    # определяем как располагается N(N_x, N_y) относительно прямых образованных
                                    # точками M1(M1_x, M1_y), M2(M2_x, M2_y), M3(M3_x, M3_y)
                                    a = (d_point[i][0] - N_x) * (d_point[k][1] - d_point[i][1]) - (d_point[k][0] - d_point[i][0]) * (d_point[i][1] - N_y)
                                    b = (d_point[k][0] - N_x) * (d_point[v][1] - d_point[k][1]) - (d_point[v][0] - d_point[k][0]) * (d_point[k][1] - N_y)
                                    c = (d_point[v][0] - N_x) * (d_point[i][1] - d_point[v][1]) - (d_point[i][0] - d_point[v][0]) * (d_point[v][1] - N_y)


                                    '''
                                    Если a, b, c имеют один и тот же знак или 0, то точка N(N_x, N_y) лежит внутри 
                                    треугольника образованного точками M1(M1_x, M1_y), M2(M2_x, M2_y), M3(M3_x, M3_y).
                                    
                                    Если a, b, c имеют различные знаки, то точка N(N_x, N_y) лежит вне треугольника 
                                    образованного точками M1(M1_x, M1_y), M2(M2_x, M2_y), M3(M3_x, M3_y).
                                    
                                    Если одно из значений a, b, c равно 0, то точка N(N_x, N_y) находится на прямой 
                                    образованной парой точек M1(M1_x, M1_y), M2(M2_x, M2_y), M3(M3_x, M3_y).
                                    
                                    Если два значения a, b, c равны 0, то точка N(N_x, N_y) лежит на пересечении прямых,
                                    образованных парами точек M1(M1_x, M1_y), M2(M2_x, M2_y), M3(M3_x, M3_y).
                                    
                                    Если все три значения a, b, c равны 0, то все три точки M1(M1_x, M1_y), 
                                    M2(M2_x, M2_y), M3(M3_x, M3_y) лежат на одной прямой, и точка N(N_x, N_y) также 
                                    лежит на этой прямой.
                                    '''
                                    # определяем положение точки N(N_x, N_y) относительно треугольника образованного
                                    # точками M1(M1_x, M1_y), M2(M2_x, M2_y), M3(M3_x, M3_y).
                                    if a >= 0 and (b >= 0 and c >= 0) or (a <= 0 and (b <= 0 and c <= 0)):
                                        '''
                                         сравнивая периметр образованный точками M1(M1_x, M1_y), M2(M2_x, M2_y), 
                                         M3(M3_x, M3_y) для точки N(N_x, N_y)  c ранее записанным в переменную
                                         Per, если условие выполняется, перезаписываем точки на новые. Так же для 
                                         отсеивания точек-дубликатов одной точки с координатами окрестности этой точки ~0.000x м. 
                                         ( не склееных в программе, не удаленных при обработке и т.д.) ставится условие
                                         perimeter > (L)
                                        '''

                                        if Per > perimeter > 0.5:
                                            # записываем координаты точки М1, и радиус от N(N_x, N_y) до M1(M1_x, M1_y)
                                            R1 = r_1
                                            M1_x = d_point[i][0]
                                            M1_y = d_point[i][1]
                                            M1_z = d_point[i][2]
                                            # записываем координаты точки М2, и радиус от N(N_x, N_y) до M2(M2_x, M2_y)
                                            R2 = r_2
                                            M2_x = d_point[k][0]
                                            M2_y = d_point[k][1]
                                            M2_z = d_point[k][2]
                                            # записываем координаты точки М3, и радиус от N(N_x, N_y) до M3(M3_x, M3_y)
                                            R3 = r_3
                                            M3_x = d_point[v][0]
                                            M3_y = d_point[v][1]
                                            M3_z = d_point[v][2]

                                            Per = perimeter  # периметр треугольника образованного точками M1, M2, M3
                                            Et = (M2_x - M1_x) * (M3_y - M1_y) - (M3_x - M1_x) * (M2_y - M1_y)
                                            # если Et=0 (a=0, b=0, c=0) то все 3 точки M1, M2, M3 лежат на одной прямой


            if Et == 0:
                # если точка N(N_x, N_y) лежит в какой то из точек M1 M2 M3
                if M1_x == N_x and M1_y == N_y:
                    u_z = round(M1_z, 6)
                    print('Точка лежит в M1')

                elif M2_x == N_x and M2_y == N_y:
                    u_z = round(M2_z, 6)
                    print('Точка лежит в M2')

                elif M3_x == N_x and M3_y == N_y:
                    u_z = round(M3_z, 6)
                    print('Точка лежит в M3')

                elif M1_y == M2_y == M3_y:  # если 3 точки лежат на одной прямой X

                    mX_min, mZ_min = -10000, None  # переменные для записи значения X, ближайшей слева точки (-10000 для входа в условие)
                    mX_max, mZ_max = 10000, None  # переменные для записи значения X, ближайшей справа точки (10000 для входа в условие)

                    for n in range(2, max_row_data + 1):  # перебираем последовательно все значения в Excel в столбце с значениями (x)
                        m_x = sheet_data.cell(row=n, column=column_X).value  # считываем и записываем в переменную значение из Excel (x)
                        m_z = sheet_data.cell(row=n, column=column_Z).value  # считываем и записываем в переменную значение из Excel (z)

                        if N_x < m_x < mX_max:  # подбираем ближайшую для N(N_x, N_y) точку справа
                            mX_max = m_x
                            mZ_max = m_z

                        elif mX_min < m_x < N_x:  # подбираем ближайшую для N(N_x, N_y) точку слева
                            mX_min = m_x
                            mZ_min = m_z

                    u_z = mZ_min + (N_x - mX_min) * ((mZ_max - mZ_min) / (mX_max - mX_min))  # определяем перемещение в N(N_x, N_y) методом интерполяции
                    round(u_z, 6)
                    print('N_x=', N_x, 'mX_min=', mX_min, 'mX_max=', mX_max, 'mZ_min=', mZ_min, 'mZ_max=', mZ_max)

                elif M1_x == M2_x == M3_x:  # если 3 точки лежат на одной прямой Y

                    mY_min, mZ_min = -10000, None  # переменные для записи значения Y, ближайшей снизу точки (-10000 для входа в условие)
                    mY_max, mZ_max = 10000, None  # переменные для записи значения Y, ближайшей сверху точки (10000 для входа в условие)

                    for n in range(2, max_row_data + 1):  # перебираем последовательно все значения в Excel в столбце с значениями (y)
                        m_y = sheet_data.cell(row=n, column=column_Y).value  # считываем и записываем в переменную значение из Excel (y)
                        m_z = sheet_data.cell(row=n, column=column_Z).value  # считываем и записываем в переменную значение из Excel (z)

                        if N_y < m_y < mY_max:  # подбираем ближайшую для N(N_x, N_y) точку сверху
                            mY_max = m_y
                            mZ_max = m_z

                        elif mY_min < m_y < N_y:  # подбираем ближайшую для N(N_x, N_y) точку снизу
                            mY_min = m_y
                            mZ_min = m_z

                    print(mY_min, mY_max, mZ_min, mZ_max)
                    u_z = mZ_min + (N_y - mY_min) * ((mZ_max - mZ_min) / (mY_max - mY_min))  # определяем перемещение в N(N_x, N_y) методом интерполяции
                    round(u_z, 6)
                    print('N_y=', N_y, 'mY_min=', mY_min, 'mY_max=', mY_max, 'mZ_min=', mZ_min, 'mZ_max=', mZ_max)

            else:  # определяем координату u_z методами аналитической геометрии (точка пересечения прямой и плоскости)
                Ax = (N_x - M1_x) * ((M2_y - M1_y) * (M3_z - M1_z) - (M3_y - M1_y) * (M2_z - M1_z))
                By = (N_y - M1_y) * ((M2_x - M1_x) * (M3_z - M1_z) - (M3_x - M1_x) * (M2_z - M1_z))
                Dz = (N_z - M1_z) * ((M2_x - M1_x) * (M3_y - M1_y) - (M3_x - M1_x) * (M2_y - M1_y))
                Et = (M2_x - M1_x) * (M3_y - M1_y) - (M3_x - M1_x) * (M2_y - M1_y)

                u_z = round(-((Ax - By + Dz) / Et), 6)  # перемещение в точке N(N_x, N_y)

            sheet_output.cell(row=1, column=x + 1, value=N_x)  # формирование шапки по длине (столбцам). Значение = значению точки N_x. (Первая строка)
            sheet_output.cell(row=y + 1, column=1, value=N_y)  # формирование шапки по высоте (строкам). Значение = значению точки N_y. (Первый столбец)
            sheet_output.cell(row=y + 1, column=x + 1, value=u_z)  # заполнение ячейки полученным значение u_z при (N_x и N_y)

            # оформление таблицы
            # оформление первой строки (по длине)
            sheet_output.cell(row=1, column=x + 1).fill = myFill
            sheet_output.cell(row=1, column=x + 1).font = myFont
            sheet_output.cell(row=1, column=x + 1).alignment = myAlignment
            sheet_output.cell(row=1, column=x + 1).border = myBorder
            # оформление первого столбца (по высоте)
            sheet_output.cell(row=y + 1, column=1).fill = myFill
            sheet_output.cell(row=y + 1, column=1).font = myFont
            sheet_output.cell(row=y + 1, column=1).alignment = myAlignment
            sheet_output.cell(row=y + 1, column=1).border = myBorder
            # оформление ячеек со значениями
            sheet_output.cell(row=y + 1, column=x + 1).border = myBorder
            sheet_output.cell(row=y + 1, column=x + 1).alignment = myAlignment

    sheet_output.cell(row=1, column=1, value='y|x')
    sheet_output.cell(row=1, column=1).fill = PatternFill(start_color='e8fdfb', end_color='e8fdfb', fill_type='solid')
    sheet_output.cell(row=1, column=1).font = myFont
    sheet_output.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    sheet_output.cell(row=1, column=1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                       top=Side(style='thin'), bottom=Side(style='thin'))

    g += 1  # переход на следующий лист
wb_output.save(file_output)  # сохранение файла с результатами
